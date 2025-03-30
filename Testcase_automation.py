import time
from typing import assert_type

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver.ie.service import Service
from selenium.webdriver.support.select import Select
import openpyxl
from openpyxl.styles import PatternFill
#need to add Selenium and openpyxl in the list of Packages of your Python Interpreter

#Prepare your Test Report excel to be updated
book = openpyxl.load_workbook("Test_Report.xlsx")
sheet = book.active
green = PatternFill(start_color="80FF00", end_color="80FF00",fill_type = "solid")
red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

def testcase_1():
    sheet.cell(row=2, column=1).value = "Testcase_1"
    driver = webdriver.Chrome()  # .Firefox()
    driver.get("https://practicesoftwaretesting.com/")
    driver.implicitly_wait(5)  # 5 seconds max timeout for all process
    driver.maximize_window()
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR,"#search-query").send_keys("ham")
    time.sleep(2)
    driver.find_element(By.XPATH,"//button[@type='submit']").click()
    time.sleep(2)
    hammer_products = driver.find_elements(By.XPATH, "//*[contains(text(), 'Hammer')]")
    hammer_number = len(hammer_products)
    if hammer_number == 7:
        sheet.cell(row=2, column=2).value = "OK"
        sheet.cell(row=2, column=3).value = ""
        sheet['B2'].fill = green
    else:
        sheet.cell(row=2, column=2).value = "NOK"
        sheet.cell(row=2, column=3).value = "Number of hammer products different than expected"
        sheet['B2'].fill = red
    driver.close()
def testcase_2():
    sheet.cell(row=3, column=1).value = "Testcase_2"
    driver = webdriver.Chrome()  # .Firefox()
    driver.get("https://practicesoftwaretesting.com/")
    driver.implicitly_wait(5)  # 5 seconds max timeout for all process
    driver.maximize_window()
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR,"#search-query").send_keys("wren")
    time.sleep(2)
    driver.find_element(By.XPATH,"//button[@type='submit']").click()
    time.sleep(2)
    wrench_products = driver.find_elements(By.XPATH, "//*[contains(text(), 'Wrench')]")
    wrench_number = len(wrench_products)
    if wrench_number == 3:
        sheet.cell(row=3, column=2).value = "OK"
        sheet.cell(row=3, column=3).value = ""
        sheet['B3'].fill = green
    else:
        sheet.cell(row=3, column=2).value = "NOK"
        sheet.cell(row=3, column=3).value = "Number of Wrench products different than expected"
        sheet['B3'].fill = red
    driver.close()
def testcase_3():
    sheet.cell(row=4, column=1).value = "Testcase_3"
    driver = webdriver.Chrome()  # .Firefox()
    driver.get("https://practicesoftwaretesting.com/")
    driver.implicitly_wait(5)  # 5 seconds max timeout for all process
    driver.maximize_window()
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR,".card").click()
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR,"#btn-add-to-cart").click()
    price_1 = driver.find_element(By.CSS_SELECTOR,"span[aria-label='unit-price']" ).text
    time.sleep(2)
    driver.back()
    driver.find_element(By.CSS_SELECTOR, "img[alt='Pliers']").click()
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, "#btn-add-to-cart").click()
    price_2 = driver.find_element(By.CSS_SELECTOR, "span[aria-label='unit-price']").text
    driver.back()
    time.sleep(10)
    driver.find_element(By.CSS_SELECTOR,".ng-fa-icon.px-1").click()
    time.sleep(10)
    cart_price = driver.find_element(By.CSS_SELECTOR, "td[data-test='cart-total']").text
    if float(cart_price[1:]) == float(price_1) + float(price_2):
        sheet.cell(row=4, column=2).value = "OK"
        sheet.cell(row=4, column=3).value = ""
        sheet['B4'].fill = green
    else:
        sheet.cell(row=4, column=2).value = "NOK"
        sheet.cell(row=4, column=3).value = "Total cart price not correct"
        sheet['B4'].fill = red
    driver.close()
def testcase_4():
    sheet.cell(row=5, column=1).value = "Testcase_4"
    driver = webdriver.Chrome()  # .Firefox()
    driver.get("https://practicesoftwaretesting.com/")
    driver.implicitly_wait(5)  # 5 seconds max timeout for all process
    driver.maximize_window()
    time.sleep(5)
    checkboxes = driver.find_elements(By.XPATH, "//input[@type='checkbox']")
    checkboxes[20].click()
    time.sleep(5)
    if checkboxes[20].get_attribute("value") == "01JQMA81Y8JXZSB12R7SGSP33Z":
        sheet.cell(row=5, column=2).value = "OK"
        sheet.cell(row=5, column=3).value = ""
        sheet['B5'].fill = green
    else:
        sheet.cell(row=5, column=2).value = "NOK"
        sheet.cell(row=5, column=3).value = "Atribute value incorrect"
        sheet['B5'].fill = red
    driver.close()
def testcase_5():
    sheet.cell(row=6, column=1).value = "Testcase_5"
    driver = webdriver.Chrome()  # .Firefox()
    driver.get("https://practicesoftwaretesting.com/")
    driver.implicitly_wait(5)  # 5 seconds max timeout for all process
    driver.maximize_window()
    action = ActionChains(driver)
    time.sleep(5)

    dropdown = Select(driver.find_element(By.CSS_SELECTOR, ".form-select"))
    dropdown.select_by_visible_text("Price (Low - High)")
    time.sleep(5)
    price_1 = driver.find_element(By.XPATH, "(//span[@class='float-end text-muted'])[1]").text
    price_2 = driver.find_element(By.XPATH, "(//span[@class='float-end text-muted'])[5]").text

    if float(price_1[1:]) <= float(price_2[1:]):
        sheet.cell(row=6, column=2).value = "OK"
        sheet.cell(row=6, column=3).value = ""
        sheet['B6'].fill = green
    else:
        sheet.cell(row=6, column=2).value = "NOK"
        sheet.cell(row=6, column=3).value = "Sort by price ascending not correct"
        sheet['B6'].fill = red
    driver.close()

testcase_1()
testcase_2()
testcase_3()
testcase_4()
testcase_5()
book.save("Test_Report.xlsx")
time.sleep(2)